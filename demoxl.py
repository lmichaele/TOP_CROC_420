import openpyxl, csv, os, datetime
from forex_python.converter import CurrencyRates
c = CurrencyRates()

directory = "G:\\Supply Chain\\Ship Files\\SHIP1175\\"

os.chdir(directory)

SHIP = 1175

f = open("SHIP1175.csv")
exampleReader = csv.reader(f)
data = list(exampleReader)
invoices = []
r=0
for row in data:
    if data[r][0] == 'IDH':
        invoices.append(data[r][5])
    r=r+1

FCV = []
r=0
for row in data:
    if data[r][0] == 'IDH':
        FCV.append(data[r][9])
    r=r+1

CI = []
r=0
for row in data:
    if data[r][0] == 'IDH':
        CIs = (float(data[r][9]))-(float(data[r][7]))
        CI.append(CIs)
    r=r+1

WH = []
r=0
for row in data:
    if data[r][1] == '8682':
        WH.append(1)
    elif data[r][1] == '8686':
        WH.append(2)
    elif data[r][1] == '8728':
        WH.append(4)
    elif data[r][1] == '8687':
        WH.append(5)
    elif data[r][1] == '18682':
        WH.append(1)
    elif data[r][1] == '18686':
        WH.append(2)
    elif data[r][1] == '18728':
        WH.append(4)
    elif data[r][1] == '18687':
        WH.append(5)
    elif data[r][1] == '8726':
        WH.append(1)
    elif data[r][1] == '8986':
        WH.append(2)
    elif data[r][1] == '8928':
        WH.append(4)
    elif data[r][1] == '8688':
        WH.append(5)
    elif data[r][1] == '18726':
        WH.append(1)
    elif data[r][1] == '18986':
        WH.append(2)
    elif data[r][1] == '18928':
        WH.append(4)
    elif data[r][1] == '18688':
        WH.append(5)
    r=r+1
        
Supplier = []
r=0
for row in data:
    if data[r][1] == '8682':
        Supplier.append(1)
    elif data[r][1] == '8686':
        Supplier.append(1)
    elif data[r][1] == '8728':
        Supplier.append(1)
    elif data[r][1] == '8687':
        Supplier.append(1)
    elif data[r][1] == '18682':
        Supplier.append(24)
    elif data[r][1] == '18686':
        Supplier.append(24)
    elif data[r][1] == '18728':
        Supplier.append(24)
    elif data[r][1] == '18687':
        Supplier.append(24)
    elif data[r][1] == '8726':
        Supplier.append(732)
    elif data[r][1] == '8986':
        Supplier.append(732)
    elif data[r][1] == '8928':
        Supplier.append(732)
    elif data[r][1] == '8688':
        Supplier.append(732)
    elif data[r][1] == '18726':
        Supplier.append(735)
    elif data[r][1] == '18986':
        Supplier.append(735)
    elif data[r][1] == '18928':
        Supplier.append(735)
    elif data[r][1] == '18688':
        Supplier.append(735)
    r=r+1

GBP = c.get_rate('GBP','AUD')
AUD = []
for i in FCV:
   AUD.append(float(i) * GBP)

XR = []
for i in AUD:
    XR.append(GBP)

SHIPLIST = []
for i in invoices:
    SHIPLIST.append(SHIP)

DMT = 'A' # get rid of me 

AS = []
for i in invoices:
    AS.append(DMT)

EDR = openpyxl.load_workbook('G:\\Supply Chain\\Data\\SHIPMENTS\\Electronic Register - SGS - Copy.xlsx')
sheet = EDR.get_sheet_by_name('Register')

nextrow = int(sheet.max_row)+1

whl = nextrow

r=0
'''
nextrow = int(sheet.max_row)+1
for i in invoices:
    sheet.cell(row=nextrow, column=11).value = int(i)
    r=r+1
    nextrow = nextrow + 1
    EDR.save('G:\\Supply Chain\\Data\\SHIPMENTS\\Electronic Register - SGS - Copy.xlsx')
'''
r=0
while r < len(invoices):
    for i in WH:
        c = EDR['Register']['G'+str(nextrow)]
        c.value = int(i)
        r=r+1
        nextrow = nextrow + 1
        
nextrow=nextrow-(len(invoices))

r=0
while r < len(invoices):
    for i in SHIPLIST:
        c = EDR['Register']['B'+str(nextrow)]
        c.value = int(i)
        r=r+1
        nextrow = nextrow + 1

nextrow=nextrow-(len(invoices))

r=0
while r < len(invoices):
    for i in invoices:
        c = EDR['Register']['K'+str(nextrow)]
        c.value = int(i)
        r=r+1
        nextrow = nextrow + 1

nextrow=nextrow-(len(invoices))

r=0
while r < len(invoices):
    for i in AS:
        c = EDR['Register']['F'+str(nextrow)]
        c.value = str(i)
        r=r+1
        nextrow = nextrow + 1

nextrow=nextrow-(len(invoices))

r=0
while r < len(invoices):
    for i in Supplier:
        c = EDR['Register']['I'+str(nextrow)]
        c.value = int(i)
        r=r+1
        nextrow = nextrow + 1

nextrow=nextrow-(len(invoices))

r=0
while r < len(invoices):
    for i in FCV:
        c = EDR['Register']['L'+str(nextrow)]
        c.value = float(i)
        r=r+1
        nextrow = nextrow + 1

nextrow=nextrow-(len(invoices))

r=0
while r < len(invoices):
    for i in CI:
        c = EDR['Register']['M'+str(nextrow)]
        c.value = float(i)
        r=r+1
        nextrow = nextrow + 1

nextrow=nextrow-(len(invoices))

r=0
while r < len(invoices):
    for i in AUD:
        c = EDR['Register']['O'+str(nextrow)] #round off
        c.value = (round(float(i)),2)
        r=r+1
        nextrow = nextrow + 1

nextrow=nextrow-(len(invoices))

r=0
while r < len(invoices):
    for i in XR:
        c = EDR['Register']['P'+str(nextrow)]
        c.value = float(i)
        r=r+1
        nextrow = nextrow + 1

nextrow=nextrow-(len(invoices))

        
EDR.save('G:\\Supply Chain\\Data\\SHIPMENTS\\Electronic Register - SGS - Copy.xlsx')




